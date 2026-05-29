import os
import requests
from requests.auth import HTTPBasicAuth
import urllib3
import json
import re
import time
import logging
import logging.handlers
import configparser
from datetime import datetime, timedelta, timezone
from datetime import date
import win32evtlogutil
import win32evtlog
import win32api
import win32con
import argparse
import subprocess
import shutil
import sys
from pathlib import Path
import jks
from cryptography import x509
from cryptography.hazmat.primitives import hashes, serialization
from cryptography.hazmat.primitives.asymmetric import rsa
from cryptography.x509.oid import NameOID
from cryptography.hazmat.primitives import serialization
from cryptography.hazmat.primitives.serialization import BestAvailableEncryption, pkcs12, Encoding, PrivateFormat, NoEncryption, BestAvailableEncryption
from cryptography.hazmat.backends import default_backend

import getpass
from openpyxl import load_workbook

import luva_lic
import luva_code
import keycloak_suite
import YD_ROLES_complement

# Pour les erreurs dans les evenements
service_base = "YD_ROLES"

contrainte_active_ini = True
contrainte_contient_luva = False # Doit faire la vérification du nom du programme : True / False
contrainte_active_lic = True # Doit faire la vérification de la license : True / False
event_active_log = True # Doit activer les logs des events : True / False
MODE_DEV = False

if contrainte_active_lic and not MODE_DEV : contrainte_active_ini = True

### Génération du nom du fichier ini
ini_filename = luva_code.get_ini_path()
if MODE_DEV : print(f"ini_filename : {ini_filename}")

### Récupération du nom du programme
nom_programme = luva_code.get_nom_programme()
if MODE_DEV : print(f"nom_programme : {nom_programme}")

### Vérification de contient service
contient_service = luva_code.nom_programme_contient(nom_programme,"SERVICE")
if contient_service :
	traitement_type = "service"
else:
	traitement_type = "exe"

### Vérification de contient LUVA
contient_luva = luva_code.nom_programme_contient(nom_programme,"LUVA")
if contrainte_contient_luva and not contient_luva :
	print(f"ERROR : Nom du programme non conforme")
	input()
	sys.exit(1)
	


### Variables complémentaires
log_event_level = "ERROR"
log_folder_level = "ERROR"
log_console_level = "WARNING"

time_sleep=0

log_levels = {
	"DEBUG": logging.DEBUG,
	"INFO": logging.INFO,
	"WARNING": logging.WARNING,
	"ERROR": logging.ERROR
}

# Désactiver les avertissements de sécurité pour les requêtes non vérifiées
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

### lecture du fichier INI pour le traitement de toutes les entrées
config = configparser.ConfigParser()

## creation des loggers
logger_service = logging.getLogger(service_base)
logger_service.setLevel(logging.WARNING)
event_handler = logging.handlers.NTEventLogHandler(service_base)
logger_service.addHandler(event_handler)

if event_active_log : luva_code.create_event_source(f"{service_base}")

### Lecture du fichier INI et gestion des logs ###
try:
	if not os.path.exists(ini_filename):
		luva_code.log_secure("0","INFO",f"Pas de fichier ini '{ini_filename}'",log_event_level,logger_service,log_console_level,traitement_type)
		log_validation = False
	else :
		luva_code.log_secure("0","INFO",f"Fichier ini '{ini_filename}' est présent",log_event_level,logger_service,log_console_level,traitement_type)
		log_validation = True
		config.read(f'{ini_filename}')
		if not config.sections():  # Vérifie si le fichier ini est vide
			if contrainte_active_ini : 
				luva_code.log_secure("0","ERROR",f"Le fichier de configuration '{ini_filename}' est vide.",log_event_level,logger_service,log_console_level,traitement_type)
				sys.exit(1)  # ferme lapp
			else :
				luva_code.log_secure("0","INFO",f"Le fichier de configuration '{ini_filename}' est vide.",log_event_level,logger_service,log_console_level,traitement_type)
except Exception as e:
	luva_code.log_secure("0","ERROR",f"Probleme de traitement du fichier '{ini_filename}': {e}",log_event_level,logger_service,log_console_level,traitement_type)
	sys.exit(1)  # ferme lapp
if MODE_DEV : print(f"log_validation : {log_validation}")
### Valide l'activation des logs
if log_validation : enable_logging = luva_code.get_enable_logging(log_validation,config,MODE_DEV)


if MODE_DEV : print(f"enable_logging : {enable_logging}")
if MODE_DEV : print(f"log_validation : {log_validation}")
if enable_logging :
	try:
		log_folder = config['logging']['log_folder']
		if MODE_DEV : print(f"log_folder : {log_folder}")
	except Exception as e: enable_logging = False
	
	try: 
		log_folder_level = config['logging']['log_folder_level']
	except Exception as e: log_folder_level = "ERROR"
	if MODE_DEV : print(f"log_folder : {log_folder}")
	if MODE_DEV : print(f"log_folder_level : {log_folder_level}")
try:
	if log_validation : 
		log_console_level = config['logging']['log_console_level']
		if MODE_DEV : print(f"log_console_level : {log_console_level}")
	else : 
		log_console_level = "INFO"
except Exception as e: log_console_level = "INFO"
if MODE_DEV : print(f"log_console_level : {log_console_level}")
try:	
	if enable_logging: os.makedirs(log_folder, exist_ok=True)  # Créer le dossier de log
except Exception as e:
	luva_code.log_secure("0","ERROR",f"creation du dossier {log_folder} impossible : {e}",log_event_level,logger_service,log_console_level,traitement_type)
	sys.exit(1)  # Quitte l'application avec un code d'erreur

try:	
	# Configure le log si activé
	if enable_logging:
		log_level = log_levels.get(log_folder_level, logging.ERROR)
		log_filename = os.path.join(log_folder, datetime.now().strftime("traitement_%Y-%m-%d.log"))
		logger_folder = logging.getLogger('folderloger')
		logger_folder.setLevel(log_level)
		file_handler = logging.FileHandler(log_filename)
		file_handler.setLevel(log_level)
		formatter = logging.Formatter('%(asctime)s - %(levelname)s - %(message)s')
		file_handler.setFormatter(formatter)
		logger_folder.addHandler(file_handler)
	else:
		logging.disable(logging.CRITICAL)  # Désactiver le logging si non activé
except Exception as e:
	luva_code.log_secure("0","ERROR",f"creation du dossier {log_folder} impossible : {e}",log_event_level,logger_service,log_console_level,traitement_type)
	sys.exit(1)  # Quitte l'application avec un code d'erreur	

try:
	log_event_level = config['logging']['log_event_level']
	luva_code.log_console("DEBUG",f"L'option 'log_event_level' est de niveau '{log_event_level}'",log_console_level,traitement_type)
except Exception as e:
	log_event_level = "ERROR"
verif = False

log_level2 = log_levels.get(log_event_level, logging.ERROR)
logger_service.setLevel(log_level2)

if contrainte_active_lic :
	licence_valide = False
	licence = luva_code.get_licence(config)
	if MODE_DEV : print(licence)
	
	try:
		licence_valide = luva_lic.valide_licence(licence,MODE_DEV)
	except Exception as e:
		licence_valide = False
	if not licence_valide :
		luva_code.log_secure("0","ERROR",f"Pas de licence VALIDE",log_event_level,logger_service,log_console_level,traitement_type)
		sys.exit(1)  # Quitte l'application avec un code d'erreur
	elif MODE_DEV :
		print(f"Licence VALIDE")

### fin traitement des logs

# Main
if __name__ == "__main__":
	
	
	try :
		KEYCLOAK_URL = YD_ROLES_complement.get_INFO_CONFIG(config,"KEYCLOAK","KEYCLOAK_URL")
		if MODE_DEV : print(KEYCLOAK_URL)
		KEYCLOAK_CLIENT = YD_ROLES_complement.get_INFO_CONFIG(config,"KEYCLOAK","KEYCLOAK_CLIENT")
		if MODE_DEV : print(KEYCLOAK_CLIENT)
		KEYCLOAK_ENV = YD_ROLES_complement.get_INFO_CONFIG(config,"KEYCLOAK","KEYCLOAK_ENV")
		if MODE_DEV : print(KEYCLOAK_ENV)
		
		REALM = f"YD-{KEYCLOAK_CLIENT}-{KEYCLOAK_ENV}"
		
		APPLICATION = YD_ROLES_complement.get_INFO_CONFIG(config,"APPLICATION","APPLICATION")
		if MODE_DEV : print(APPLICATION)
		EXCEL = YD_ROLES_complement.get_INFO_CONFIG(config,"APPLICATION","EXCEL")
		if MODE_DEV : print(EXCEL)
		FEUILLE_EXCEL = YD_ROLES_complement.get_INFO_CONFIG(config,"APPLICATION","FEUILLE_EXCEL")
		if MODE_DEV : print(FEUILLE_EXCEL)
		COLONNE_ROLE = YD_ROLES_complement.get_INFO_CONFIG(config,"APPLICATION","COLONNE_ROLE")
		if MODE_DEV : print(COLONNE_ROLE)
		COLONNE_TENANT = YD_ROLES_complement.get_INFO_CONFIG(config,"APPLICATION","COLONNE_TENANT")
		if MODE_DEV : print(COLONNE_TENANT)
		DEBUT_CELLULE = YD_ROLES_complement.get_INFO_CONFIG(config,"APPLICATION","DEBUT_CELLULE")
		if MODE_DEV : print(DEBUT_CELLULE)
		FIN_CELLULLE = YD_ROLES_complement.get_INFO_CONFIG(config,"APPLICATION","FIN_CELLULLE")
		if MODE_DEV : print(FIN_CELLULLE)
		ACTION = YD_ROLES_complement.get_INFO_CONFIG(config,"APPLICATION","ACTION")
		if MODE_DEV : print(ACTION)
		
	except Exception as e:
		luva_code.log_console("ERROR",f"Récuperation des info du INI : {e}",log_console_level,traitement_type)
		sys.exit(1)
	
	log_texte = f"Information du fichier ini recupéré avec succés" 
	luva_code.log_event("1","DEBUG",log_texte,log_event_level,log_console_level,log_folder_level,logger_folder,logger_service,traitement_type,enable_logging)
		
	
	try :
		REALM = YD_ROLES_complement.get_INFO_CONFIG(config,"LUVA","REALM").strip()
		if REALM == "" or REALM == "ERROR" : REALM = f"YD-{KEYCLOAK_CLIENT}-{KEYCLOAK_ENV}"
		if MODE_DEV : print(REALM)
	except Exception as e:
		if MODE_DEV : print("pas de luva REALM")
		REALM = f"YD-{KEYCLOAK_CLIENT}-{KEYCLOAK_ENV}"
	try :
		CLIENT_NAME = YD_ROLES_complement.get_INFO_CONFIG(config,"LUVA","CLIENT").strip()
		if CLIENT_NAME == "" or CLIENT_NAME == "ERROR": CLIENT_NAME = f"YDG-{KEYCLOAK_CLIENT}-{KEYCLOAK_ENV}"
		if MODE_DEV : print(CLIENT_NAME)
	except Exception as e:
		if MODE_DEV : print("pas de luva CLIENT")
		CLIENT_NAME = f"YDG-{KEYCLOAK_CLIENT}-{KEYCLOAK_ENV}"
	try :
		USERNAME = YD_ROLES_complement.get_INFO_CONFIG(config,"LUVA","USERNAME").strip()
		if USERNAME == "ERROR" : USERNAME = ""
		if MODE_DEV : print(USERNAME)
	except Exception as e:
		USERNAME = ""
		if MODE_DEV : print("pas de luva USERNAME")
	try :
		PASSWORD = YD_ROLES_complement.get_INFO_CONFIG(config,"LUVA","PASSWORD").strip()
		if PASSWORD == "ERROR" : PASSWORD = ""
		if MODE_DEV : print(PASSWORD)
	except Exception as e:
		PASSWORD = ""
		if MODE_DEV : print("pas de luva PASSWORD")
	try :
		MODE_DEV = YD_ROLES_complement.get_BOOL_CONFIG(config,"LUVA","MODE_DEV")
		if MODE_DEV == "ERROR" : MODE_DEV = False
		if MODE_DEV : print(MODE_DEV)
	except Exception as e:
		if MODE_DEV : print("pas de MODE_DEV")
	
	#Variable pour tout le traitement
	KEYCLOAK_BASE = KEYCLOAK_URL
	if USERNAME == "" :
		while True:
			USERNAME = input("Entrez le nom utilisateur de connexion : ").strip()
			if USERNAME:
				break
	
	if PASSWORD == "" :
		while True:
			PASSWORD = input("Entrez le mot de passe de connexion : ").strip()
			if PASSWORD:
				break
	
	if MODE_DEV : print(f"Connexion > {KEYCLOAK_BASE}")
	if MODE_DEV : print(f"sur le realm > {REALM}")
	if MODE_DEV : print(f"sur le client > {CLIENT_NAME}")
	if MODE_DEV : print(f"sur le user > {USERNAME}")
	if MODE_DEV : print(f"sur le password > {PASSWORD}")
	
	TIMEOUT_TIME = 10
	VERIF_SSL = False
	
	log_texte = f"Connexion au keycloak {KEYCLOAK_BASE} sur le realm {REALM} le client {CLIENT_NAME}" 
	luva_code.log_event("1","DEBUG",log_texte,log_event_level,log_console_level,log_folder_level,logger_folder,logger_service,traitement_type,enable_logging)
		
	
	# token_gestion_return = keycloak_suite.keycloak_master_barear(KEYCLOAK_URL, "master", "admin-cli", "password", USERNAME, PASSWORD ,TIMEOUT_TIME , VERIF_SSL)
	# token_gestion_error = token_gestion_return["error"]
	
	# if token_gestion_error != "":
		# log_texte = f"{token_gestion_return["code"]}{token_gestion_return["error"]} : {token_gestion_return["message2"]}{token_gestion_return["message"]}" 
		# log_event("1","ERROR",log_texte,log_event_level,log_console_level,logger_folder,logger_service,traitement_type)
	# else :
		# token_gestion_result = token_gestion_return["result"]
		
	# token_gestion_access_token = token_gestion_result.get("access_token")
	# token_gestion_refresh_token = token_gestion_result.get("refresh_token")
	
	# token_gestion_expires_in = token_gestion_result.get("expires_in")
	# token_gestion_refresh_expires_in = token_gestion_result.get("refresh_expires_in")
	
	# token_gestion_token_type = token_gestion_result.get("token_type")
	# token_gestion_scope = token_gestion_result.get("scope")
	
	
	CLIENT_ID = "admin-cli"
	GRANT_TYPE = "password"
	CLIENT_SECRET = None
	
	USERNAME = "administrateur"
	PASSWORD = "toor"
	VERIF_SSL = False
	TIMEOUT_TIME = 10
	
	if MODE_DEV : print(f"Récuperation du token du realm {REALM} avec le user {USERNAME}")
	
	
	token_gestion_return = keycloak_suite.keycloak_master_barear(KEYCLOAK_URL, "master", "admin-cli", "password", USERNAME, PASSWORD ,TIMEOUT_TIME , VERIF_SSL)
	token_gestion_error = token_gestion_return["error"]
	if token_gestion_error != "":
		log_texte = f"{token_gestion_return["code"]}{token_gestion_return["error"]} : {token_gestion_return["message2"]}{token_gestion_return["message"]}" 
		luva_code.log_event("1","ERROR",log_texte,log_event_level,log_console_level,log_folder_level,logger_folder,logger_service,traitement_type,enable_logging)
		sys.exit(1)
	else :
		token_gestion_result = token_gestion_return["result"]
	
	token_gestion_access_token = token_gestion_result.get("access_token")
	token_gestion_refresh_token = token_gestion_result.get("refresh_token")
	
	token_gestion_expires_in = token_gestion_result.get("expires_in")
	token_gestion_refresh_expires_in = token_gestion_result.get("refresh_expires_in")
	
	token_gestion_token_type = token_gestion_result.get("token_type")
	token_gestion_scope = token_gestion_result.get("scope")
	
	log_texte= f"recuperation du token de type {token_gestion_token_type} réalisé avec succès"
	if MODE_DEV : print(log_texte)
	luva_code.log_event("1","INFO",log_texte,log_event_level,log_console_level,log_folder_level,logger_folder,logger_service,traitement_type,enable_logging)
	
	if MODE_DEV : print("--------------")
	if MODE_DEV : print(f"Recup Client ID")
	if MODE_DEV : print(f" ")
	client_list_return = keycloak_suite.keycloak_client_id(KEYCLOAK_URL,token_gestion_access_token,REALM,CLIENT_NAME, VERIF_SSL,MODE_DEV)
	
	client_list_return_error = client_list_return["error"]
	client_list_return_result = client_list_return["result"]
	client_list_return_message = client_list_return["message"]
	if client_list_return_error != "":
		luva_code.log_console("ERROR",f"{client_list_return["code"]}{client_list_return["error"]} : {client_list_return["message2"]}{client_list_return["message"]}",log_console_level,traitement_type)
	else :
		CLIENT_ID = client_list_return["message2"]
		if MODE_DEV : print(f"Client id  = {CLIENT_ID}")
	
	log_texte= f"recuperation id client {CLIENT_ID} réalisé avec succès"
	if MODE_DEV : print(log_texte)
	luva_code.log_event("1","INFO",log_texte,log_event_level,log_console_level,log_folder_level,logger_folder,logger_service,traitement_type,enable_logging)
		
	try:
		MATRICE_ROLE = EXCEL.strip()
		FEUILLE_LISTE_ROLE = FEUILLE_EXCEL.strip()
		APPLICATION = APPLICATION.strip()
		DEBUT_CELLULE_ROLE = int(DEBUT_CELLULE)
		FIN_CELLULLE_ROLE = int(FIN_CELLULLE)
	except Exception as e:
		luva_code.log_console("ERROR",f"Verification de la demande : {e}",log_console_level,traitement_type)
		sys.exit(1)
	
	ID_APP = ""
	if APPLICATION == "GESTION" : ID_APP = "ydg"
	if APPLICATION == "ANALYSE" : ID_APP = "yda"
	if ID_APP == "" : 
		luva_code.log_console("ERROR",f"L'application doit être 'GESTION' ou 'ANALYSE' et non '{APPLICATION}'",log_console_level,traitement_type)
		sys.exit(1)
	try:
		wb = load_workbook(filename=MATRICE_ROLE, data_only=True)
		ws = wb[f"{FEUILLE_LISTE_ROLE}"] if FEUILLE_LISTE_ROLE else wb.active
	except Exception as e:
		luva_code.log_console("ERROR",f"Lecture du fichier matrice : {e}",log_console_level,traitement_type)
		sys.exit(1)
		
	ID_ACT = ""
	if ACTION == "CREATION" : ID_ACT = "create"
	if ACTION == "SUPPRESSION" : ID_ACT = "suppr"
	if ID_ACT == "" : 
		luva_code.log_console("ERROR",f"L'action doit être 'SUPPRESSION' ou 'CREATION' et non '{ACTION}'",log_console_level,traitement_type)
		sys.exit(1)
	if MODE_DEV : print(REALM)
	log_texte = f"Lancement du traitement"
	luva_code.log_event("1","INFO",log_texte,log_event_level,log_console_level,log_folder_level,logger_folder,logger_service,traitement_type,enable_logging)
	for i in range(DEBUT_CELLULE_ROLE, (FIN_CELLULLE_ROLE + 1)):  # i va de 0 à 9, incrémentation de 1 par défaut
		ROLE_GROUPE = ws[f"{COLONNE_ROLE}{i}"].value
		ROLE_TENANT = ws[f"{COLONNE_TENANT}{i}"].value
		log_texte = f"valeur COLONNE{i} : {ID_APP}_{ROLE_TENANT}_{ROLE_GROUPE}"
		if MODE_DEV : print(f"{log_texte}")
		luva_code.log_event("1","DEBUG",log_texte,log_event_level,log_console_level,log_folder_level,logger_folder,logger_service,traitement_type,enable_logging)
	
		ROLE_NAME = f"{ID_APP}_{ROLE_TENANT}_{ROLE_GROUPE}"
		ROLE_DESCRIPTION = f"{ID_APP}_{ROLE_TENANT}_{ROLE_GROUPE}"
		
		if ID_ACT == "create" :
			client_list_return = keycloak_suite.keycloak_create_role(	KEYCLOAK_URL,
																		token_gestion_access_token,
																		REALM,CLIENT_ID,
																		ROLE_NAME,ROLE_DESCRIPTION,
																		VERIF_SSL,
																		MODE_DEV)
		
		if ID_ACT == "suppr" :
			client_list_return =  keycloak_suite.keycloak_delete_role(	KEYCLOAK_BASE,
																		token_gestion_access_token, 
																		REALM,CLIENT_ID,
																		ROLE_NAME,ROLE_DESCRIPTION,
																		VERIF_SSL,
																		MODE_DEV)
		
		
		client_list_return_error = client_list_return["error"]
		client_list_return_result = client_list_return["result"]
		client_list_return_message = client_list_return["message"]
		if client_list_return_error != "":
			log_texte = f"{client_list_return["code"]}{client_list_return["error"]} : {client_list_return["message2"]}{client_list_return["message"]}"
			luva_code.log_event("1","INFO",log_texte,log_event_level,log_console_level,log_folder_level,logger_folder,logger_service,traitement_type,enable_logging)
		else :
			log_texte = f"{client_list_return_message}"
			luva_code.log_event("1","INFO",log_texte,log_event_level,log_console_level,log_folder_level,logger_folder,logger_service,traitement_type,enable_logging)
	print("Fin du traitement")
	input()