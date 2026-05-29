import os
import requests
from requests.auth import HTTPBasicAuth
import urllib3
import json
import re
import time
import logging
import logging.handlers
import configparser
from datetime import datetime, timedelta, timezone
from datetime import date
import win32evtlogutil
import win32evtlog
import win32api
import win32con
import argparse
import subprocess
import shutil
import sys
from pathlib import Path
import jks
from cryptography import x509
from cryptography.hazmat.primitives import hashes, serialization
from cryptography.hazmat.primitives.asymmetric import rsa
from cryptography.x509.oid import NameOID
from cryptography.hazmat.primitives import serialization
from cryptography.hazmat.primitives.serialization import BestAvailableEncryption, pkcs12, Encoding, PrivateFormat, NoEncryption, BestAvailableEncryption
from cryptography.hazmat.backends import default_backend

from openpyxl import load_workbook

import luva_lic
import token_gestion
import client_user
import group_gestion

import keycloak_suite

# Pour les erreurs dans les evenements
service_base = "Luva_cert"

# Pour le traitement des services (vérification du nom services dans le programme)

programme = os.path.splitext(os.path.basename(sys.argv[0]))[0]
path = Path(sys.executable).resolve() if getattr(__import__("sys"), "frozen", False) else Path(__file__).resolve()
exe_filename = f"{programme}.exe"
terme_service = "service"
contient_service = terme_service.lower() in exe_filename.lower()

if contient_service :
	traitement_type = "service"
else:
	traitement_type = "exe"

log_event_level = "ERROR"
log_folder_level = "ERROR"
log_console_level = "INFO"

time_sleep=0

log_levels = {
	"DEBUG": logging.DEBUG,
	"INFO": logging.INFO,
	"WARNING": logging.WARNING,
	"ERROR": logging.ERROR
}

### affiche les logs dans les Evenements windows ###
def log_service(level,log_text):
	if level == "DEBUG" and log_event_level == "DEBUG" :
		logger_service.debug(f"{log_text}")
	if level == "INFO" and (log_event_level == "DEBUG" or log_event_level == "INFO") :
		logger_service.info(f"{log_text}")
	elif level == "WARNING" and (log_event_level == "DEBUG" or log_event_level == "INFO" or log_event_level == "WARNING"):
		logger_service.warning(f"{log_text}")
	elif level == "ERROR" :
		logger_service.error(f"{log_text}")

### affiche les logs dans la console ###
def log_console(level,log_text):
	if traitement_type == "exe" :
		current_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
		if level == "DEBUG" and log_console_level == "DEBUG" :
			print(f"{current_time} > {level} > {log_text}")
		if level == "INFO" and (log_console_level == "DEBUG" or log_console_level == "INFO") :
			print(f"{current_time} > {level} > {log_text}")
		elif level == "WARNING" and (log_console_level == "DEBUG" or log_console_level == "INFO" or log_console_level == "WARNING"):
			print(f"{current_time} > {level} > {log_text}")
		elif level == "ERROR" :
			print(f"{current_time} > {level} > {log_text}")
	
### affiche les logs de façon securisé avant l'initialisations des paramètres (event et console) ###
def log_secure(type,level,log_text):
	if type == "0" :
		log_service(level,log_text)
	log_console(level,log_text)
	if traitement_type == "exe" and level == "ERROR" :
		input("Appuyez sur une touche pour quitter...")

### création d'une source d'événements si elle n'existe pas pour le mode service ###
def create_event_source(source_name):
	try:
		win32evtlogutil.ReportEvent(source_name, 1, eventType=win32evtlog.EVENTLOG_INFORMATION_TYPE, strings=[source_name])
	except Exception as e:
		print("error")
		win32api.RegCreateKey(win32con.HKEY_LOCAL_MACHINE, f'SYSTEM\\CurrentControlSet\\Services\\EventLog\\Application\\{source_name}')
		win32api.RegSetValue(win32api.RegOpenKey(win32con.HKEY_LOCAL_MACHINE, f'SYSTEM\\CurrentControlSet\\Services\\EventLog\\Application\\{source_name}'), 'EventMessageFile', win32con.REG_SZ, os.path.abspath(__file__))
		win32api.RegSetValue(win32api.RegOpenKey(win32con.HKEY_LOCAL_MACHINE, f'SYSTEM\\CurrentControlSet\\Services\\EventLog\\Application\\{source_name}'), 'TypesSupported', win32con.REG_DWORD, 7)

### enregistre les logs dans un fichier ###
def log_enreg(level,log_text):
	current_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
	if level == "DEBUG" and log_folder_level == "DEBUG" :
		logger_folder.debug(f"{log_text}")
	if level == "INFO" and (log_folder_level == "DEBUG" or log_folder_level == "INFO") :
		logger_folder.info(f"{log_text}")
	elif level == "WARNING" and (log_folder_level == "DEBUG" or log_folder_level == "INFO" or log_folder_level == "WARNING"):
		logger_folder.warning(f"{log_text}")
	elif level == "ERROR" :
		logger_folder.error(f"{log_text}")
		
### traitement des logs en fonction du type ###
### 2 > pour la console ###
### 1 > pour log dans un fichier et la console ###
### 0 > pour les event win, log dans un fichier et la console ###
def log_event(type,level,log_text):
	if type == "0" :
		if enable_logging :
			log_enreg(level,log_text)
		log_service(level,log_text)
		log_console(level,log_text)
	elif type == "1" :
		if enable_logging :
			log_enreg(level,log_text)
		log_console(level,log_text)
	elif type == "2" :
		log_console(level,log_text)

# Désactiver les avertissements de sécurité pour les requêtes non vérifiées
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)
		
### lecture du fichier INI et traitement de toutes les entrées ###
config = configparser.ConfigParser()
logger_service = logging.getLogger(service_base)
logger_service.setLevel(logging.WARNING)
event_handler = logging.handlers.NTEventLogHandler(service_base)
logger_service.addHandler(event_handler)
create_event_source(f"{service_base}")

### Recupéaration des du nom du programme ###
programme = os.path.splitext(os.path.basename(sys.argv[0]))[0]
path = Path(sys.executable).resolve() if getattr(__import__("sys"), "frozen", False) else Path(__file__).resolve()

ini_filename = f"{programme}.ini"



### Sécurisation des variables de logs ###
try:
	
	if not os.path.exists(ini_filename):
		log_secure("0","INFO",f"Pas de fichier ini '{ini_filename}'")
		log_validation = False
	else :
		log_secure("0","INFO",f"Fichier ini '{ini_filename}' est présent")
		log_validation = True
		config.read(f'{ini_filename}')
		if not config.sections():  # Vérifie si le fichier ini est vide
			raise FileNotFoundError(f"Le fichier de configuration '{ini_filename}' est vide.")
except Exception as e:
	log_secure("0","ERROR",f"Probleme de traitement du fichier '{ini_filename}': {e}")
	sys.exit(1)  # ferme lapp

try:
	if log_validation :
		enable_logging = config.getboolean('logging', 'enable_logging')
	else :
		enable_logging = False
except Exception as e:
	enable_logging = False

if enable_logging :
	try:
		log_folder = config['logging']['log_folder']
	except Exception as e:
		log_secure("0","ERROR",f"L'option 'log_folder' est manquante dans le fichier 'config.ini': {e}")
		sys.exit(1)  # Quitte l'application avec un code d'erreur
	try:
		log_folder_level = config['logging']['log_folder_level']
	except Exception as e:
		log_folder_level = "ERROR"

try:
	if log_validation :
		log_console_level = config['logging']['log_console_level']
	else :
		log_console_level = "INFO"
except Exception as e:
	log_console_level = "INFO"

try:	
	if enable_logging:
		os.makedirs(log_folder, exist_ok=True)  # Créer le dossier de log
except Exception as e:
	log_secure("0","ERROR",f"creation du dossier {log_folder} impossible : {e}")
	sys.exit(1)  # Quitte l'application avec un code d'erreur

try:	
	# Configure le log si activé
	if enable_logging:
		log_level = log_levels.get(log_folder_level, logging.ERROR)
		log_filename = os.path.join(log_folder, datetime.now().strftime("traitement_%Y-%m-%d.log"))
		logger_folder = logging.getLogger('folderloger')
		logger_folder.setLevel(log_level)
		file_handler = logging.FileHandler(log_filename)
		file_handler.setLevel(log_level)
		formatter = logging.Formatter('%(asctime)s - %(levelname)s - %(message)s')
		file_handler.setFormatter(formatter)
		logger_folder.addHandler(file_handler)
	else:
		logging.disable(logging.CRITICAL)  # Désactiver le logging si non activé
except Exception as e:
	log_secure("0","ERROR",f"creation du dossier {log_folder} impossible : {e}")
	sys.exit(1)  # Quitte l'application avec un code d'erreur	

try:
	log_event_level = config['logging']['log_event_level']
	log_console("DEBUG",f"L'option 'log_event_level' est de niveau '{log_event_level}'")
except Exception as e:
	log_event_level = "ERROR"
verif = False


log_level2 = log_levels.get(log_event_level, logging.ERROR)
logger_service.setLevel(log_level2)

# Main
if __name__ == "__main__":
	
	
	print(f"Récuperation du token du realm")
	
	KEYCLOAK_BASE = "https://ydgluva24q:8281"   # URL de Keycloak (sans slash final)
	REALM = "master"
	CLIENT_ID = "admin-cli"
	GRANT_TYPE = "password"
	CLIENT_SECRET = None
	USERNAME = "administrateur"
	PASSWORD = "toor"
	VERIF_SSL = False
	TIMEOUT_TIME = 10
	
	print(f"Récuperation du token du realm {REALM} avec le user {USERNAME}")
	
	token_gestion_return = keycloak_suite.keycloak_master_barear(KEYCLOAK_BASE, REALM, CLIENT_ID, GRANT_TYPE, USERNAME, PASSWORD ,TIMEOUT_TIME , VERIF_SSL)
	token_gestion_error = token_gestion_return["error"]
	
	if token_gestion_error != "":
		log_console("ERROR",f"{token_gestion_return["code"]}{token_gestion_return["error"]} : {token_gestion_return["message2"]}{token_gestion_return["message"]}")
	else :
		token_gestion_result = token_gestion_return["result"]
		
	token_gestion_access_token = token_gestion_result.get("access_token")
	token_gestion_refresh_token = token_gestion_result.get("refresh_token")
	
	token_gestion_expires_in = token_gestion_result.get("expires_in")
	token_gestion_refresh_expires_in = token_gestion_result.get("refresh_expires_in")
	
	token_gestion_token_type = token_gestion_result.get("token_type")
	token_gestion_scope = token_gestion_result.get("scope")
	
	
	print(f"Récuperation du token du realm de type > {token_gestion_token_type} ")
	# print(token_gestion_access_token)
	# print(token_gestion_token_type)
	# print(token_gestion_scope)
	
	
	print(f"---------------------------------------")
	print(f"Vérifier si le client existe")
	
	KEYCLOAK_BASE = "https://ydgluva24q:8281"   # URL de Keycloak (sans slash final)
	REALM = "YD-LAINX-VM"
	CLIENT_ID = "admin-cli"
	GRANT_TYPE = "password"
	CLIENT_SECRET = None
	USERNAME = "administrateur"
	PASSWORD = "toor"
	VERIF_SSL = False
	TIMEOUT_TIME = 10
	MODE_DEV = True

	
	print("--------------")
	print(f"Client liste")
	print(f" ")
	client_list_return = keycloak_suite.keycloak_client_liste(KEYCLOAK_BASE,token_gestion_access_token,REALM, VERIF_SSL,MODE_DEV)
	client_list_return_error = client_list_return["error"]
	client_list_return_result = client_list_return["result"]
	client_list_return_message = client_list_return["message"]
	if client_list_return_error != "":
		log_console("ERROR",f"{client_list_return["code"]}{client_list_return["error"]} : {client_list_return["message2"]}{client_list_return["message"]}")
	else :
		print(f"{client_list_return_message}")
	
	CLIENT = "YDG-LAINX-VM"
	print("--------------")
	print(f"Client ID")
	print(f" ")
	client_list_return = keycloak_suite.keycloak_client_id(KEYCLOAK_BASE,token_gestion_access_token,REALM,CLIENT, VERIF_SSL,MODE_DEV)
	
	client_list_return_error = client_list_return["error"]
	client_list_return_result = client_list_return["result"]
	client_list_return_message = client_list_return["message"]
	if client_list_return_error != "":
		log_console("ERROR",f"{client_list_return["code"]}{client_list_return["error"]} : {client_list_return["message2"]}{client_list_return["message"]}")
	else :
		print(f"{client_list_return_message}")
		CLIENT_ID = client_list_return["message2"]
		
	MATRICE_ROLE = r"D:\PYTHON\Youdoc_Keysuit\YoudocFichier.xlsx"
	FEUILLE_LISTE_ROLE = "Feuil1"
	APPLICATION = "ydg"
	COLONNE_ROLE = "A"
	COLONNE_ID = "A"
	COLONNE_TENANT = "C"
	DEBUT_CELLULE_ROLE = "22"
	FIN_CELLULLE_ROLE = "61"
	
	MATRICE_SECU = r"D:\PYTHON\Youdoc_Keysuit\YoudocFichier.xlsx"
	DEBUT_LISTE_ROLE = "A22"
	FIN_LISTE_ROLE = "A61"
	SHEET_NAME = "Feuil1"
	
	
	# Remplacez 'votre_fichier.xlsx' par le chemin de votre fichier Excel
	
	
	
	wb = load_workbook(filename=MATRICE_ROLE, data_only=True)
	ws = wb[f"{FEUILLE_LISTE_ROLE}"] if FEUILLE_LISTE_ROLE else wb.active
	
	
	CLIENT = "YDG-LAINX-VM"
	
	
	ACTION = "CREATION"
	
	for i in range(int(DEBUT_CELLULE_ROLE), int(FIN_CELLULLE_ROLE)):  # i va de 0 à 9, incrémentation de 1 par défaut
		ROLE_GROUPE = ws[f"{COLONNE_ROLE}{i}"].value
		ROLE_TENANT = ws[f"{COLONNE_TENANT}{i}"].value
		print(f"valeur COLONNE{i} : {APPLICATION}_{ROLE_TENANT}_{ROLE_GROUPE}")
		ROLE_NAME = f"{APPLICATION}_{ROLE_TENANT}_{ROLE_GROUPE}"
		ROLE_DESCRIPTION = f"{APPLICATION}_{ROLE_TENANT}_{ROLE_GROUPE}"
		
		if ETAPE == "CREATE" :
			client_list_return = keycloak_suite.keycloak_create_role(KEYCLOAK_BASE,token_gestion_access_token,REALM,CLIENT_ID,ROLE_NAME,ROLE_DESCRIPTION,VERIF_SSL,MODE_DEV)
		
		if ETAPE == "DELETE" :
			client_list_return =  keycloak_suite.keycloak_delete_role(	KEYCLOAK_BASE,
							token_gestion_access_token, 
							REALM,CLIENT_ID,
							ROLE_NAME,ROLE_DESCRIPTION,
							VERIF_SSL,
							MODE_DEV)
		
		
		client_list_return_error = client_list_return["error"]
		client_list_return_result = client_list_return["result"]
		client_list_return_message = client_list_return["message"]
		if client_list_return_error != "":
			log_console("ERROR",f"{client_list_return["code"]}{client_list_return["error"]} : {client_list_return["message2"]}{client_list_return["message"]}")
		else :
			print(f"{client_list_return_message}")
	sys.exit(1)
		
	print(CLIENT_ID)
	CLIENT = "YDG-LAINX-VM"
	ROLE_NAME = valeur
	ROLE_DESCRIPTION = valeur
	
	print("--------------")
	print(f"keycloak_create_role")
	print(f" ")
	client_list_return = keycloak_suite.keycloak_create_role(KEYCLOAK_BASE,token_gestion_access_token,REALM,CLIENT_ID,ROLE_NAME,ROLE_DESCRIPTION,VERIF_SSL,MODE_DEV)
	
	client_list_return_error = client_list_return["error"]
	client_list_return_result = client_list_return["result"]
	client_list_return_message = client_list_return["message"]
	if client_list_return_error != "":
		log_console("ERROR",f"{client_list_return["code"]}{client_list_return["error"]} : {client_list_return["message2"]}{client_list_return["message"]}")
	else :
		print(f"{client_list_return_message}")
		
	sys.exit(1)
	
	
	
	
	
	
	
	
	
	
	
	
	
	print("--------------")
	print(f"Recuperation du token BAREAR")
	
	KEYCLOAK_BASE = "https://ydgluva24q:8281"   # URL de Keycloak (sans slash final)
	REALM = "YD-LAINX-VM"
	CLIENT_ID = "YDG-LAINX-VM"
	GRANT_TYPE = "password"
	CLIENT_SECRET = "k42xJ7DrVT6lW9l8nDorRtc64mL6l76c"  # Si le client est public, laissez None
	USERNAME = "lainxluva"
	PASSWORD = "passwordT2I"
	VERIF_SSL = False
	TIMEOUT_TIME = 10
	
	token_client_return = token_gestion.keycloak_token_client(KEYCLOAK_BASE, REALM, CLIENT_ID, GRANT_TYPE, USERNAME, PASSWORD, CLIENT_SECRET ,TIMEOUT_TIME , VERIF_SSL)
	token_gestion_error = token_client_return["error"]
	
	if token_gestion_error != "":
		log_console("ERROR",f"{token_client_return["code"]}{token_client_return["error"]} : {token_client_return["message2"]}{token_client_return["message"]}")
	else :
		token_client_result = token_client_return["result"]
		
	token_client_access_token = token_client_result.get("access_token")
	token_client_refresh_token = token_client_result.get("refresh_token")
	
	token_client_expires_in = token_client_result.get("expires_in")
	token_client_refresh_expires_in = token_client_result.get("refresh_expires_in")
	
	token_client_token_type = token_client_result.get("token_type")
	token_client_scope = token_client_result.get("scope")
	
	print(token_client_access_token)
	print(token_client_token_type)
	print(token_client_scope)
	
	print("--------------")
	print(f"decode")
	token_gestion.token_decode(token_client_access_token)
	
	
	sys.exit(1)
	print("--------------")
	print(f"Token Realm")
	
	KEYCLOAK_BASE = "https://ydgluva24q:8281"   # URL de Keycloak (sans slash final)
	REALM = "master"
	CLIENT_ID = "admin-cli"
	GRANT_TYPE = "password"
	CLIENT_SECRET = None
	USERNAME = "administrateur"
	PASSWORD = "toor"
	VERIF_SSL = False
	TIMEOUT_TIME = 10
	
	token_gestion_return = token_gestion.token_gestion_realm(KEYCLOAK_BASE, REALM, CLIENT_ID, GRANT_TYPE, USERNAME, PASSWORD ,TIMEOUT_TIME , VERIF_SSL)
	token_gestion_error = token_gestion_return["error"]
	
	if token_gestion_error != "":
		log_console("ERROR",f"{token_gestion_return["code"]}{token_gestion_return["error"]} : {token_gestion_return["message2"]}{token_gestion_return["message"]}")
	else :
		token_gestion_result = token_gestion_return["result"]
		
	token_gestion_access_token = token_gestion_result.get("access_token")
	token_gestion_refresh_token = token_gestion_result.get("refresh_token")
	
	token_gestion_expires_in = token_gestion_result.get("expires_in")
	token_gestion_refresh_expires_in = token_gestion_result.get("refresh_expires_in")
	
	token_gestion_token_type = token_gestion_result.get("token_type")
	token_gestion_scope = token_gestion_result.get("scope")
	
	print(token_gestion_access_token)
	print(token_gestion_token_type)
	print(token_gestion_scope)
	
	
	print("--------------")
	print(f"decode")
	token_gestion.token_decode(token_gestion_access_token)
	
	REALM = "YD-LAINX-VM"
	USER_ID = "JDOE6"
	USER_MAIL = "TEST@TEST6.FR"
	USER_FIRSTNAME = "JON"
	USER_LASTNAME = "DOE"
	USER_ENABLED = True
	USER_VERIFIED = True
	USER_PASSWORD = "passwordT2I"
	USER_TEMPORARY = False
	
	print("--------------")
	print(f"User creation")
	client_user_return = client_user.client_user_creation(token_gestion_access_token,KEYCLOAK_BASE, REALM, USER_ID, USER_MAIL, USER_FIRSTNAME, USER_LASTNAME, USER_ENABLED, USER_VERIFIED, USER_PASSWORD, USER_TEMPORARY , VERIF_SSL)
	client_user_return_error = client_user_return["error"]
	client_user_result = client_user_return["result"]
	client_user_message = client_user_return["message"]
	if client_user_return_error != "":
		log_console("ERROR",f"{client_user_return["code"]}{client_user_return["error"]} : {client_user_return["message2"]}{client_user_return["message"]}")
	else :
		print(f"{client_user_message}")
		
	print("--------------")
	print(f"User id")
	client_user_return = client_user.client_user_id(token_gestion_access_token,KEYCLOAK_BASE, REALM, USER_ID, VERIF_SSL)
	client_user_return_error = client_user_return["error"]
	client_user_result = client_user_return["result"]
	client_user_message = client_user_return["message"]
	if client_user_return_error != "":
		log_console("ERROR",f"{client_user_return["code"]}{client_user_return["error"]} : {client_user_return["message2"]}{client_user_return["message"]}")
	else :
		print(f"{client_user_message}")
	
	print("--------------")
	print(f"User liste")
	client_user_return = client_user.client_user_liste(token_gestion_access_token,KEYCLOAK_BASE, REALM, VERIF_SSL)
	client_user_return_error = client_user_return["error"]
	client_user_result = client_user_return["result"]
	client_user_message = client_user_return["message"]
	if client_user_return_error != "":
		log_console("ERROR",f"{client_user_return["code"]}{client_user_return["error"]} : {client_user_return["message2"]}{client_user_return["message"]}")
	else :
		print(f"{client_user_message}")

	
	print("--------------")
	print(f"Groupe lecture")
	YDG_BASE = "https://ydgluva24q:8443"   # URL de Keycloak (sans slash final)
	TENANT = "001"
	USER_ID = "lainxluva"
	GROUP = "001-adm-grp"
	VERIF_SSL = False
	group_gestion.ydg_group_creation(token_client_access_token,YDG_BASE, TENANT,USER_ID,GROUP,VERIF_SSL)
	# group_lecture_return = group_gestion.ydg_group_creation(token_client_access_token,YDG_BASE, TENANT,USER_ID,GROUP,VERIF_SSL)
	# group_lecture_return_error = group_lecture_return["error"]
	# group_lecture_result = group_lecture_return["result"]
	# group_lecture_message = group_lecture_return["message"]
	# if group_lecture_return_error != "":
		# log_console("ERROR",f"{group_lecture_return["code"]}{group_lecture_return["error"]} : {group_lecture_return["message2"]}{group_lecture_return["message"]}")
	# else :
		# print(f"{group_lecture_message}")
	
	print("Fin")
	input()