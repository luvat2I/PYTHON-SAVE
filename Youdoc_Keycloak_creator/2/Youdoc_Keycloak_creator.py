import os
import requests
from requests.auth import HTTPBasicAuth
import urllib3
import ast, json
import re
import time
import ast
import logging
import logging.handlers
import configparser
from datetime import datetime, timedelta, timezone
from datetime import date
import win32evtlogutil
import win32evtlog
import win32api
import win32con
import argparse
import subprocess
import shutil
import sys
from pathlib import Path
import jks
from cryptography import x509
from cryptography.hazmat.primitives import hashes, serialization
from cryptography.hazmat.primitives.asymmetric import rsa
from cryptography.x509.oid import NameOID
from cryptography.hazmat.primitives import serialization
from cryptography.hazmat.primitives.serialization import BestAvailableEncryption, pkcs12, Encoding, PrivateFormat, NoEncryption, BestAvailableEncryption
from cryptography.hazmat.backends import default_backend
import getpass

from openpyxl import load_workbook

import luva_lic
import luva_code

import keycloak_suite
import group_gestion
import Youdoc_Keycloak_creator_complement

# Pour les erreurs dans les evenements
service_base = "Youdoc_Keycloak_creator"
licence_base = "9736234122426398669" # sur l'année 2026

contrainte_active_ini = True #doit faire la vérification du fichier .ini
contrainte_contient_luva = False # Doit faire la vérification du nom du programme : True / False
contrainte_active_lic = True # Doit faire la vérification de la license : True / False
event_active_log = False # Doit activer les logs des events : True / False
MODE_DEV = False

if contrainte_active_lic and not MODE_DEV : contrainte_active_ini = True #secu pour le mode dev

### Génération du nom du fichier ini
ini_filename = luva_code.get_ini_path()
if MODE_DEV : print(f"> DEV > ini_filename : {ini_filename}")

### Récupération du nom du programme
nom_programme = luva_code.get_nom_programme()
if MODE_DEV : print(f"> DEV > nom_programme : {nom_programme}")

### Vérification de contient service
contient_service = luva_code.nom_programme_contient(nom_programme,"SERVICE")
if contient_service :
    traitement_type = "service"
else:
    traitement_type = "exe"

### Vérification de contient LUVA
contient_luva = luva_code.nom_programme_contient(nom_programme,"LUVA")
if contrainte_contient_luva and not contient_luva :
    print(f"ERROR : Nom du programme non conforme")
    input()
    sys.exit(1)

### Variables complémentaires
log_event_level = "ERROR"
log_folder_level = "ERROR"
log_console_level = "WARNING"

time_sleep=0

log_levels = {"DEBUG": logging.DEBUG,"INFO": logging.INFO,"WARNING": logging.WARNING,"ERROR": logging.ERROR}

# Désactiver les avertissements de sécurité pour les requêtes non vérifiées
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

### lecture du fichier INI pour le traitement de toutes les entrées
config = configparser.ConfigParser()

## creation des loggers
logger_service = logging.getLogger(service_base)
logger_service.setLevel(logging.WARNING)
event_handler = logging.handlers.NTEventLogHandler(service_base)
logger_service.addHandler(event_handler)

if event_active_log : luva_code.create_event_source(f"{service_base}")

### Lecture du fichier INI et gestion des logs ###
try:
    if not os.path.exists(ini_filename):
        luva_code.log_secure("0","INFO",f"Pas de fichier ini '{ini_filename}'",log_event_level,logger_service,log_console_level,traitement_type)
        log_validation = False
    else :
        luva_code.log_secure("0","INFO",f"Fichier ini '{ini_filename}' est présent",log_event_level,logger_service,log_console_level,traitement_type)
        log_validation = True
        config.read(f'{ini_filename}')
        if not config.sections():  # Vérifie si le fichier ini est vide
            if contrainte_active_ini : 
                luva_code.log_secure("0","ERROR",f"Le fichier de configuration '{ini_filename}' est vide.",log_event_level,logger_service,log_console_level,traitement_type)
                sys.exit(1)  # ferme lapp
            else :
                luva_code.log_secure("0","INFO",f"Le fichier de configuration '{ini_filename}' est vide.",log_event_level,logger_service,log_console_level,traitement_type)
except Exception as e:
    luva_code.log_secure("0","ERROR",f"Probleme de traitement du fichier '{ini_filename}': {e}",log_event_level,logger_service,log_console_level,traitement_type)
    sys.exit(1)  # ferme lapp
enable_logging = False
### Valide l'activation des logs
if log_validation : enable_logging = luva_code.get_enable_logging(log_validation,config)

if enable_logging :
    try:
        log_folder = config['logging']['log_folder']
    except Exception as e: enable_logging = False
    
    try: 
        log_folder_level = config['logging']['log_folder_level']
    except Exception as e: log_folder_level = "ERROR"

try:
    if log_validation : log_console_level = config['logging']['log_console_level']
    else : log_console_level = "INFO"
except Exception as e: log_console_level = "INFO"

try:    
    if enable_logging: os.makedirs(log_folder, exist_ok=True)  # Créer le dossier de log
except Exception as e:
    luva_code.log_secure("0","ERROR",f"creation du dossier {log_folder} impossible : {e}",log_event_level,logger_service,log_console_level,traitement_type)
    sys.exit(1)  # Quitte l'EXCEL_APPLICATION avec un code d'erreur

try:    
    # Configure le log si activé
    if enable_logging:
        log_level = log_levels.get(log_folder_level, logging.ERROR)
        log_filename = os.path.join(log_folder, datetime.now().strftime("traitement_%Y-%m-%d.log"))
        logger_folder = logging.getLogger('folderloger')
        logger_folder.setLevel(log_level)
        file_handler = logging.FileHandler(log_filename)
        file_handler.setLevel(log_level)
        formatter = logging.Formatter('%(asctime)s - %(levelname)s - %(message)s')
        file_handler.setFormatter(formatter)
        logger_folder.addHandler(file_handler)
    else:
        logging.disable(logging.CRITICAL)  # Désactiver le logging si non activé
except Exception as e:
    luva_code.log_secure("0","ERROR",f"creation du dossier {log_folder} impossible : {e}",log_event_level,logger_service,log_console_level,traitement_type)
    input()
    sys.exit(1)  # Quitte l'EXCEL_APPLICATION avec un code d'erreur    

try:
    log_event_level = config['logging']['log_event_level']
    luva_code.log_console("DEBUG",f"L'option 'log_event_level' est de niveau '{log_event_level}'",log_console_level,traitement_type)
except Exception as e:
    log_event_level = "ERROR"
verif = False

log_level2 = log_levels.get(log_event_level, logging.ERROR)
logger_service.setLevel(log_level2)

if contrainte_active_ini :
    licence_valide = False
    
    licence = luva_code.get_licence(config)
    if licence_base and licence == "ERROR" :
        licence = licence_base
    
    if MODE_DEV : print(f"> DEV > licence > {licence}")
    
    LUVA_VERIF_LICENCE = luva_code.get_param(config,'LUVA','LICENCE')
    if not LUVA_VERIF_LICENCE["error"] :
        licence = LUVA_VERIF_LICENCE["return"]
    
    try:
        licence_valide = luva_lic.valide_licence(licence,MODE_DEV)
    except Exception as e:
        licence_valide = False
        
    if not licence_valide :
        luva_code.log_secure("0","ERROR",f"ERREUR PYTHON INCONNU",log_event_level,logger_service,log_console_level,traitement_type)
        input()
        sys.exit(1)  # Quitte l'EXCEL_APPLICATION avec un code d'erreur
    elif MODE_DEV :
        print(f"> DEV > licence > VALIDE")

### fin traitement des logs

# Main
if __name__ == "__main__":
    
    # ---------------------------------------------------------------------------------------
    # Lecture INI FILE
    # ---------------------------------------------------------------------------------------
    
    print("")
    print(f">   > Début du traitement du programme de création automatique des Roles Keycloak pour YDG")
    print("")
    print(f"> 1 > -----------------------------------------------------------------------")
    print(f"> 1 > = Récuperation des paramètres du fichier {ini_filename}")
    
    
    # ++++++++++++++++++++++++++++++
    # Recup variables ENVIRONNEMENT
    # ++++++++++++++++++++++++++++++
    
    CLIENT_PARAM = Youdoc_Keycloak_creator_complement.get_param(config,'ENVIRONNEMENT','CLIENT')
    if CLIENT_PARAM["error"] :
        print(CLIENT_PARAM["return"])
        print("Appuyer sur entrer pour fermer")
        input()
        sys.exit(1)
    else :
        CLIENT = CLIENT_PARAM["return"]
    
    ENVIRONNEMENT_PARAM = Youdoc_Keycloak_creator_complement.get_param(config,'ENVIRONNEMENT','ENVIRONNEMENT')
    if ENVIRONNEMENT_PARAM["error"] :
        print(ENVIRONNEMENT_PARAM["return"])
        print("Appuyer sur entrer pour fermer")
        input()
        sys.exit(1)
    else :
        ENVIRONNEMENT = ENVIRONNEMENT_PARAM["return"]
    
    
    
    
    GENERATION_RAPPORT = False
    
    
    
    # ++++++++++++++++++++++++++++++
    # Recup variables KEYCLOAK
    # ++++++++++++++++++++++++++++++
    
    KEYCLOAK_URL_PARAM = Youdoc_Keycloak_creator_complement.get_param(config,'KEYCLOAK','KEYCLOAK_URL')
    if KEYCLOAK_URL_PARAM["error"] :
        print(KEYCLOAK_URL_PARAM["return"])
        print("Appuyer sur entrer pour fermer")
        input()
        sys.exit(1)
    else :
        KEYCLOAK_URL = KEYCLOAK_URL_PARAM["return"]
    
    KEYCLOAK_LOGIN_PARAM = Youdoc_Keycloak_creator_complement.get_param(config,'KEYCLOAK','KEYCLOAK_LOGIN')
    if KEYCLOAK_LOGIN_PARAM["error"] :
        print(KEYCLOAK_LOGIN_PARAM["return"])
        print("Appuyer sur entrer pour fermer")
        input()
        sys.exit(1)
    else :
        KEYCLOAK_LOGIN = KEYCLOAK_LOGIN_PARAM["return"]
    
    KEYCLOAK_PASSWORD_PARAM = Youdoc_Keycloak_creator_complement.get_param(config,'KEYCLOAK','KEYCLOAK_PASSWORD')
    if KEYCLOAK_PASSWORD_PARAM["error"] :
        print(KEYCLOAK_PASSWORD_PARAM["return"])
        print("Appuyer sur entrer pour fermer")
        input()
        sys.exit(1)
    else :
        KEYCLOAK_PASSWORD = KEYCLOAK_PASSWORD_PARAM["return"]
    
    KEYCLOAK_CLIENTSECRET_PARAM = Youdoc_Keycloak_creator_complement.get_param(config,'KEYCLOAK','KEYCLOAK_CLIENTSECRET')
    if KEYCLOAK_CLIENTSECRET_PARAM["error"] :
        print(KEYCLOAK_CLIENTSECRET_PARAM["return"])
        print("Appuyer sur entrer pour fermer")
        input()
        sys.exit(1)
    else :
        KEYCLOAK_CLIENTSECRET = KEYCLOAK_CLIENTSECRET_PARAM["return"]
    
    # ++++++++++++++++++++++++++++++
    # Recup variables KEYCLOAK_PARAM
    # ++++++++++++++++++++++++++++++
    
    KEYCLOAK_TIMEOUT_TIME_PARAM = Youdoc_Keycloak_creator_complement.get_param(config,'KEYCLOAK_PARAM','KEYCLOAK_TIMEOUT_TIME')
    if KEYCLOAK_TIMEOUT_TIME_PARAM["error"] :
        KEYCLOAK_TIMEOUT_TIME = 10
    else :
        KEYCLOAK_TIMEOUT_TIME = int(KEYCLOAK_TIMEOUT_TIME_PARAM["return"])
    
    KEYCLOAK_VERIF_SSL_PARAM = Youdoc_Keycloak_creator_complement.get_bool_param(config,'KEYCLOAK_PARAM','KEYCLOAK_VERIF_SSL')
    if KEYCLOAK_VERIF_SSL_PARAM["error"] :
        KEYCLOAK_VERIF_SSL = False
    else :
        KEYCLOAK_VERIF_SSL = KEYCLOAK_VERIF_SSL_PARAM["return"]
    
    KEYCLOAK_CREATE_ROLE_PARAM = Youdoc_Keycloak_creator_complement.get_bool_param(config,'KEYCLOAK_PARAM','KEYCLOAK_CREATE_ROLE')
    if KEYCLOAK_CREATE_ROLE_PARAM["error"] :
        KEYCLOAK_CREATE_ROLE = False
    else :
        KEYCLOAK_CREATE_ROLE = KEYCLOAK_CREATE_ROLE_PARAM["return"]
    
    
    GENERATION_RAPPORT_ROLE_PARAM = Youdoc_Keycloak_creator_complement.get_bool_param(config,'KEYCLOAK_PARAM','GENERATION_RAPPORT')
    if GENERATION_RAPPORT_ROLE_PARAM["error"] :
        GENERATION_RAPPORT_ROLE = False
    else :
        GENERATION_RAPPORT_ROLE = GENERATION_RAPPORT_ROLE_PARAM["return"]

    # ++++++++++++++++++++++++++++++
    # Init des variables KEYCLOAK
    # ++++++++++++++++++++++++++++++
    
    KEYCLOAK_REALM = "master"
    CALCUL_CLIENT_ID = "admin-cli"
    KEYCLOAK_GRANT_TYPE = "password"
    
    
    # ++++++++++++++++++++++++++++++
    # Recup variables EXCEL
    # ++++++++++++++++++++++++++++++
    
    EXCEL_FICHIER_PARAM = Youdoc_Keycloak_creator_complement.get_param(config,'EXCEL','EXCEL_FICHIER')
    if EXCEL_FICHIER_PARAM["error"] :
        print(EXCEL_FICHIER_PARAM["return"])
        print("Appuyer sur entrer pour fermer")
        input()
        sys.exit(1)
    else :
        EXCEL_FICHIER = EXCEL_FICHIER_PARAM["return"]
    
    
    # ++++++++++++++++++++++++++++++
    # Recup variables YOUDOC
    # ++++++++++++++++++++++++++++++
    
    YOUDOC_URL_PARAM = Youdoc_Keycloak_creator_complement.get_param(config,'YOUDOC','YOUDOC_URL')
    if YOUDOC_URL_PARAM["error"] :
        print(YOUDOC_URL_PARAM["return"])
        print("Appuyer sur entrer pour fermer")
        input()
        sys.exit(1)
    else :
        YOUDOC_URL = YOUDOC_URL_PARAM["return"]
    
    
    YOUDOC_TENANT_PARAM = Youdoc_Keycloak_creator_complement.get_param(config,'YOUDOC','YOUDOC_TENANT')
    if YOUDOC_TENANT_PARAM["error"] :
        print(YOUDOC_TENANT_PARAM["return"])
        print("Appuyer sur entrer pour fermer")
        input()
        sys.exit(1)
    else :
        YOUDOC_TENANT = YOUDOC_TENANT_PARAM["return"]
    
    
    YOUDOC_LOGIN_PARAM = Youdoc_Keycloak_creator_complement.get_param(config,'YOUDOC','YOUDOC_LOGIN')
    if YOUDOC_LOGIN_PARAM["error"] :
        print(YOUDOC_LOGIN_PARAM["return"])
        print("Appuyer sur entrer pour fermer")
        input()
        sys.exit(1)
    else :
        YOUDOC_LOGIN = YOUDOC_LOGIN_PARAM["return"]
    
    YOUDOC_PASSWORD_PARAM = Youdoc_Keycloak_creator_complement.get_param(config,'YOUDOC','YOUDOC_PASSWORD')
    if YOUDOC_PASSWORD_PARAM["error"] :
        print(YOUDOC_PASSWORD_PARAM["return"])
        print("Appuyer sur entrer pour fermer")
        input()
        sys.exit(1)
    else :
        YOUDOC_PASSWORD = YOUDOC_PASSWORD_PARAM["return"]
     
    # ++++++++++++++++++++++++++++++
    # Recup variables YOUDOC_PARAM
    # ++++++++++++++++++++++++++++++
    
    YOUDOC_TIMEOUT_TIME_PARAM = Youdoc_Keycloak_creator_complement.get_param(config,'KEYCLOAK_PARAM','YOUDOC_TIMEOUT_TIME')
    if YOUDOC_TIMEOUT_TIME_PARAM["error"] :
        YOUDOC_TIMEOUT_TIME = 10
    else :
        YOUDOC_TIMEOUT_TIME = int(YOUDOC_TIMEOUT_TIME_PARAM["return"])
    
    YOUDOC_VERIF_SSL_PARAM = Youdoc_Keycloak_creator_complement.get_bool_param(config,'KEYCLOAK_PARAM','YOUDOC_VERIF_SSL')
    if YOUDOC_VERIF_SSL_PARAM["error"] :
        YOUDOC_VERIF_SSL = False
    else :
        YOUDOC_VERIF_SSL = YOUDOC_VERIF_SSL_PARAM["return"]
    
    YOUDOC_CREATE_GROUPE_PARAM = Youdoc_Keycloak_creator_complement.get_bool_param(config,'YOUDOC_PARAM','YOUDOC_CREATE_GROUPE')
    if YOUDOC_CREATE_GROUPE_PARAM["error"] :
        YOUDOC_CREATE_GROUPE = False
    else :
        YOUDOC_CREATE_GROUPE = YOUDOC_CREATE_GROUPE_PARAM["return"]
    
    GENERATION_RAPPORT_GROUPE_PARAM = Youdoc_Keycloak_creator_complement.get_bool_param(config,'YOUDOC_PARAM','GENERATION_RAPPORT')
    if GENERATION_RAPPORT_GROUPE_PARAM["error"] :
        GENERATION_RAPPORT_GROUPE = False
    else :
        GENERATION_RAPPORT_GROUPE = GENERATION_RAPPORT_GROUPE_PARAM["return"]
        
    # ++++++++++++++++++++++++++++++
    # Recup des variables IDP
    # ++++++++++++++++++++++++++++++
    
    CREATE_MAPPER_PARAM = Youdoc_Keycloak_creator_complement.get_bool_param(config,'IDP','CREATE_MAPPER')
    if CREATE_MAPPER_PARAM["error"] :
        CREATE_MAPPER = False
    else :
        CREATE_MAPPER = CREATE_MAPPER_PARAM["return"]
    
    GENERATION_RAPPORT_MAPPER_PARAM = Youdoc_Keycloak_creator_complement.get_bool_param(config,'IDP','GENERATION_RAPPORT')
    if GENERATION_RAPPORT_MAPPER_PARAM["error"] :
        GENERATION_RAPPORT_MAPPER = False
    else :
        GENERATION_RAPPORT_MAPPER = GENERATION_RAPPORT_MAPPER_PARAM["return"]
    
    # ++++++++++++++++++++++++++++++
    # Init des variables LUVA
    # ++++++++++++++++++++++++++++++
    
    
    KEYCLOAK_VERIF_DEBUG_PARAM = Youdoc_Keycloak_creator_complement.get_bool_param(config,'LUVA','DEBUG')
    if KEYCLOAK_VERIF_DEBUG_PARAM["error"] :
        DEBUG = False
    else :
        DEBUG = KEYCLOAK_VERIF_DEBUG_PARAM["return"]
    
    KEYCLOAK_VERIF_DEBUG_PARAM = Youdoc_Keycloak_creator_complement.get_bool_param(config,'LUVA','IS_DEBUG')
    if KEYCLOAK_VERIF_DEBUG_PARAM["error"] :
        is_debug = False
    else :
        is_debug = KEYCLOAK_VERIF_DEBUG_PARAM["return"]
    
    EXCEL_FIRST_PARAM = Youdoc_Keycloak_creator_complement.get_param(config,'LUVA','EXCEL_FIRST')
    if EXCEL_FIRST_PARAM["error"] :
        EXCEL_FIRST = "14"
    else :
        EXCEL_FIRST = EXCEL_FIRST_PARAM["return"]
    
    EXCEL_APPLICATION_YDG_PARAM = Youdoc_Keycloak_creator_complement.get_param(config,'LUVA','EXCEL_APPLICATION')
    if EXCEL_APPLICATION_YDG_PARAM["error"] :
        EXCEL_APPLICATION = "ydg"
    else :
        EXCEL_APPLICATION = EXCEL_APPLICATION_YDG_PARAM["return"]
    
    EXCEL_FEUILLE_YDG_PARAM = Youdoc_Keycloak_creator_complement.get_param(config,'LUVA','EXCEL_FEUILLE_YDG')
    if EXCEL_FEUILLE_YDG_PARAM["error"] :
        EXCEL_FEUILLE_YDG = "YDG_ENTRA_ID"
    else :
        EXCEL_FEUILLE_YDG = EXCEL_FEUILLE_YDG_PARAM["return"]
    
    CALCUL_CLIENT_REALM_PARAM = Youdoc_Keycloak_creator_complement.get_param(config,'LUVA','CALCUL_CLIENT_REALM')
    if CALCUL_CLIENT_REALM_PARAM["error"] :
        CALCUL_CLIENT_REALM = f"YD-{CLIENT}-{ENVIRONNEMENT}"
    else :
        CALCUL_CLIENT_REALM = CALCUL_CLIENT_REALM_PARAM["return"]
    
    CALCUL_CLIENT_PARAM = Youdoc_Keycloak_creator_complement.get_param(config,'LUVA','CALCUL_CLIENT')
    if CALCUL_CLIENT_PARAM["error"] :
        CALCUL_CLIENT = f"YDG-{CLIENT}-{ENVIRONNEMENT}"
    else :
        CALCUL_CLIENT = CALCUL_CLIENT_PARAM["return"]
    
    KEYCLOAK_ACTION_PARAM = Youdoc_Keycloak_creator_complement.get_param(config,'LUVA','ACTION')
    if KEYCLOAK_ACTION_PARAM["error"] :
        ACTION = "CREATION"
    else :
        ACTION = KEYCLOAK_ACTION_PARAM["return"]
    
    IDP_ALIAS_PARAM = Youdoc_Keycloak_creator_complement.get_param(config,'LUVA','IDP_ALIAS')
    if IDP_ALIAS_PARAM["error"] :
        IDP_ALIAS = f"YD-{CLIENT}-{ENVIRONNEMENT}"
    else :
        IDP_ALIAS = IDP_ALIAS_PARAM["return"]
        
    KEYCLOAK_REALM = CALCUL_CLIENT_REALM
    
    if GENERATION_RAPPORT_ROLE or GENERATION_RAPPORT_GROUPE or GENERATION_RAPPORT_MAPPER :
        GENERATION_RAPPORT = True
    else :
        GENERATION_RAPPORT = False
    
    
    
    
    
    
    
    
    
    
    
    if is_debug :
        print(f"> 1 > ----------------------------------------------------")
        print(f"> 1 > ++ IS_DEBUG = {is_debug}")
        print(f"> 1 > ++ CLIENT_REALM = {CALCUL_CLIENT_REALM}")
        print(f"> 1 > ++ CLIENT = {CALCUL_CLIENT}")
        print(f"> 1 > ++ EXCEL_FEUILLE_YDG = {EXCEL_FEUILLE_YDG}")
        print(f"> 1 > ++ EXCEL_APPLICATION = {EXCEL_APPLICATION}")
        print(f"> 1 > ++ EXCEL_FIRST = {EXCEL_FIRST}")
        print(f"> 1 > ++ ACTION = {ACTION} (CREATION / SUPPRESSION)")
    
    print(f"> 1 > -----------------------------------------------------------------------")
    print(f"> 1 > | KEYCLOAK |")
    print(f"> 1 > + URL Keycloak = {KEYCLOAK_URL}")
    print(f"> 1 > + REALM = {CALCUL_CLIENT_REALM}")
    print(f"> 1 > + CLIENT ID = {CALCUL_CLIENT}")
    print(f"> 1 > + CLIENT SECRET = {KEYCLOAK_CLIENTSECRET}")
    print(f"> 1 > + LOGIN Keycloak = {KEYCLOAK_LOGIN}")
    print(f"> 1 > + PASSWORD Keycloak ={KEYCLOAK_PASSWORD}")
    print(f"> 1 > + TIMEOUT Keycloak ={KEYCLOAK_TIMEOUT_TIME}")
    print(f"> 1 > + SSL Keycloak = {KEYCLOAK_VERIF_SSL}")

    
    print(f"> 1 > -----------------------------------------------------------------------")
    print(f"> 1 > | YOUDOC |")
    print(f"> 1 > + URL Youdoc = {YOUDOC_URL}")
    print(f"> 1 > + TENANT Youdoc = {YOUDOC_TENANT}")
    print(f"> 1 > + LOGIN Youdoc = {YOUDOC_LOGIN}")
    print(f"> 1 > + PASSWORD Youdoc ={YOUDOC_PASSWORD}")
    print(f"> 1 > + TIMEOUT Youdoc ={YOUDOC_TIMEOUT_TIME}")
    print(f"> 1 > + SSL Youdoc = {YOUDOC_VERIF_SSL}")
    
    print(f"> 1 > -----------------------------------------------------------------------")
    print(f"> 1 > | IDP |")
    if CREATE_MAPPER :
        print(f"> 1 > + ALIAS IDP = {IDP_ALIAS}")
    else :
        print(f"> 1 > + PAS DE CREATION DE MAPPER")
    print(f"> 1 > -----------------------------------------------------------------------")
    print(f"> 1 > = FIN > Récuperation des paramètres")
    print(f"> 1 > -----------------------------------------------------------------------")
    print("")
    
    
    
    
    print(f"> 2 > -----------------------------------------------------------------------")
    print(f"> 2 > = Connexions Keycloak et Youdoc")
    print(f"> 2 > -----------------------------------------------------------------------")
    
    print(f"> 2 > -----------------------------------------------------------------------")
    print(f"> 2 > | Connexion KEYCLOAK |")
    
    YOUDOC_CONNEXION = group_gestion.YdgGroupClass(YOUDOC_URL, YOUDOC_TENANT, YOUDOC_LOGIN, YOUDOC_PASSWORD, YOUDOC_TIMEOUT_TIME, YOUDOC_VERIF_SSL,
    KEYCLOAK_URL, CALCUL_CLIENT_ID, KEYCLOAK_CLIENTSECRET, CALCUL_CLIENT_REALM,KEYCLOAK_TIMEOUT_TIME, KEYCLOAK_VERIF_SSL, is_debug)
    
    YOUDOC_TOKEN = YOUDOC_CONNEXION.recup_barear()
    if YOUDOC_TOKEN["error"] :
        print(YOUDOC_TOKEN["return"])
        print("Appuyer sur entrer pour fermer")
        input()
        sys.exit(1)
    else :
        print(f"> 2 > + Récupération du token et connexion > Connexion OK")
        print(f"> 2 > + Récupération du token et connexion > Token OK")
        if DEBUG : 
            print(f"")
            print(f"> 2 > Token :")
            print(f"{YOUDOC_TOKEN["return"]}")
            print(f"")
    print(f"> 2 > + Récupération du token et connexion > FIN")
    print("")
    
    print(f"> 2 > -----------------------------------------------------------------------")
    print(f"> 2 > | Connexion YOUDOC |")
    KEYCLOAK_CONNECT = keycloak_suite.Keycloak_youdoc(KEYCLOAK_URL, KEYCLOAK_REALM,CALCUL_CLIENT_REALM, CALCUL_CLIENT_ID, KEYCLOAK_GRANT_TYPE, KEYCLOAK_LOGIN, KEYCLOAK_PASSWORD ,KEYCLOAK_TIMEOUT_TIME , KEYCLOAK_VERIF_SSL, is_debug)
    
    KEYCLOAK_TOKEN = KEYCLOAK_CONNECT.recup_token_barear()
    if KEYCLOAK_TOKEN["error"] :
        print(KEYCLOAK_TOKEN["return"])
        print("Appuyer sur entrer pour fermer")
        input()
        sys.exit(1)
    else :
        print(f"> 2 > Récupération du token et connexion > Connexion OK")
        print(f"> 2 > Récupération du token et connexion > Token OK")
        if DEBUG : 
            print(f"")
            print(f"> 2 > Token :")
            print(f"{KEYCLOAK_TOKEN["return"]}")
            print(f"")
    print(f"> 2 > Récupération du token et connexion > FIN")
    
    print(f"> 2 > -----------------------------------------------------------------------")
    print(f"> 2 > = FIN > Connexions Keycloak et Youdoc")
    print(f"> 2 > -----------------------------------------------------------------------")
    print("")
    
    
    
    
    
    
    
    
    print(f"> 3 > -----------------------------------------------------------------------")
    if CREATE_MAPPER :
        print(f"> 3 > = Validation CLIENT et IDENTITY PROVIDER")
    else :
        print(f"> 3 > = Validation CLIENT")
    print(f"> 3 > -----------------------------------------------------------------------")
    
    print(f"> 3 > -----------------------------------------------------------------------")
    print(f"> 3 > | Validation CLIENT |")
    
    KEYCLOAK_ID_CLIENT = KEYCLOAK_CONNECT.recup_client_id(CALCUL_CLIENT)
    if KEYCLOAK_ID_CLIENT["error"] :
        print(KEYCLOAK_ID_CLIENT["return"])
        print("Appuyer sur entrer pour fermer")
        input()
        sys.exit(1)
    else :
        print(f"> 3 > Le client {CALCUL_CLIENT} existe")
        if DEBUG : 
            print(f"> 3 > ID client :")
            print(KEYCLOAK_ID_CLIENT["return"])
        print(f"> 3 > Connexion au client {CALCUL_CLIENT} > FINI")
    
    if CREATE_MAPPER : 
        print(f"> 3 > -----------------------------------------------------------------------")
        print(f"> 3 > | Validation IDENTITY PROVIDER |")
        IDP_VALIDE = False
        KEYCLOAK_ID_PROVIDER = KEYCLOAK_CONNECT.recup_identity_provider(IDP_ALIAS)
        if KEYCLOAK_ID_PROVIDER["error"] :
            print(KEYCLOAK_ID_PROVIDER["return"])
            print("Appuyer sur entrer pour fermer")
            input()
            sys.exit(1)
        else :
            print(f"> 3 > L'IDP {IDP_ALIAS} existe")
            if DEBUG : 
                print(f"> 3 > IDP :")
                print(KEYCLOAK_ID_PROVIDER["return"])
            IDP_VALIDE = True
            print(f"> 3 > Connexion à l'IDP {IDP_ALIAS} > FINI")
    
    
    
    print(f"> 3 > -----------------------------------------------------------------------")
    if CREATE_MAPPER :
        print(f"> 3 > = FIN > Validation CLIENT et IDENTITY PROVIDER")
    else :
        print(f"> 3 > = FIN > Validation CLIENT")
    print(f"> 3 > -----------------------------------------------------------------------")
    
    
    print(f"> 4 > -----------------------------------------------------------------------")
    if CREATE_MAPPER :
        print(f"> 4 > = Modification des GROUPES, des ROLES et des MAPPERS")
    else :
        print(f"> 4 > = Modification des GROUPES et des ROLES")
    print(f"> 4 > -----------------------------------------------------------------------")
    print(f"> 4 > -----------------------------------------------------------------------")
    print(f"> 4 > traitement excel : {EXCEL_FICHIER}")
    
    if CREATE_MAPPER or YOUDOC_CREATE_GROUPE or KEYCLOAK_CREATE_ROLE :
        workbook = load_workbook(filename=EXCEL_FICHIER, data_only=True)
        sheet  = workbook[f"{EXCEL_FEUILLE_YDG}"]
        
        derniere_ligne_groupe = 0
        for row in range(1, sheet.max_row + 1):
            if sheet.cell(row=row, column=1).value is not None:
                derniere_ligne_groupe = row
                
        if is_debug : print(f"derniere_ligne_groupe = {derniere_ligne_groupe}")
        
        derniere_ligne_id = 0
        for row in range(1, sheet.max_row + 1):
            if sheet.cell(row=row, column=2).value is not None:
                derniere_ligne_id = row
        if is_debug : print(f"derniere_ligne_groupe = {derniere_ligne_groupe}")
        derniere_ligne_tenant = 0
        for row in range(1, sheet.max_row + 1):
            if sheet.cell(row=row, column=3).value is not None:
                derniere_ligne_tenant = row
        if is_debug : print(f"derniere_ligne_groupe = {derniere_ligne_groupe}")
        if derniere_ligne_groupe != derniere_ligne_groupe and derniere_ligne_groupe != derniere_ligne_groupe :
            print("ERREUR > Il n'y a pas toutes les informations des groupes, des id et des tenants dans le fichiers xlsx il faut corriger le fichier")
            print("Appuyer sur entrer pour fermer")
            input()
            sys.exit(1)
        derniere_ligne_groupe = derniere_ligne_groupe + 1
        nb_role = int(derniere_ligne_groupe) - int(EXCEL_FIRST)
        
        print(f"> 4 > Modification de {nb_role} entrées ")
        
        for row in range(14, derniere_ligne_groupe):
            if sheet.cell(row=row, column=1).value is not None:
                
                
                
                try :
                    GROUPE_NAME = f"{sheet.cell(row=row, column=1).value}"
                    ROLE_NAME = f"{EXCEL_APPLICATION}_{sheet.cell(row=row, column=3).value}_{sheet.cell(row=row, column=1).value}"
                    MAPPER_NAME = f"{sheet.cell(row=row, column=1).value}"
                    MAPPER_ID = f"{sheet.cell(row=row, column=2).value}"
                    #verif token avant creation
                    YOUDOC_TOKEN_EXPIRE = YOUDOC_CONNEXION.token_valide()
                    if YOUDOC_TOKEN_EXPIRE["error"] :
                        print(YOUDOC_TOKEN_EXPIRE["return"])
                        YOUDOC_TOKEN_EXPIRE_token = YOUDOC_TOKEN_EXPIRE["return"]
                    else :
                        YOUDOC_TOKEN_EXPIRE_token = YOUDOC_TOKEN_EXPIRE["return"]
                    if DEBUG : print(f"> 4 > Token YD : {YOUDOC_TOKEN_EXPIRE_token}")
                    if YOUDOC_TOKEN_EXPIRE_token :
                        print(f"> 4 >             > TOKEN YD expiré ")
                        YOUDOC_TOKEN = YOUDOC_CONNEXION.recup_barear()
                        if YOUDOC_TOKEN["error"] :
                            print(YOUDOC_TOKEN["return"])
                            print("Appuyer sur entrer pour fermer")
                            input()
                            sys.exit(1)
                        else :
                            if DEBUG :
                                print(f"> 4 > Token : {YOUDOC_TOKEN["return"]}")
                        print(f"> 4 >             > TOKEN YD renouvellé ")
                        
                    KEYCLOAK_TOKEN_EXPIRE = KEYCLOAK_CONNECT.token_valide()
                    if KEYCLOAK_TOKEN_EXPIRE["error"] :
                        print(KEYCLOAK_TOKEN_EXPIRE["return"])
                        KEYCLOAK_TOKEN_EXPIRE_token = KEYCLOAK_TOKEN_EXPIRE["return"]
                    else :
                        KEYCLOAK_TOKEN_EXPIRE_token = KEYCLOAK_TOKEN_EXPIRE["return"]
                    if DEBUG : print(f"> 4 > Token KC : {KEYCLOAK_TOKEN_EXPIRE}")
                    if KEYCLOAK_TOKEN_EXPIRE_token :
                        print(f"> 4 >             > TOKEN KC expiré ")
                        KEYCLOAK_TOKEN = KEYCLOAK_CONNECT.recup_barear()
                        if KEYCLOAK_TOKEN["error"] :
                            print(KEYCLOAK_TOKEN["return"])
                            print("Appuyer sur entrer pour fermer")
                            input()
                            sys.exit(1)
                        else :
                            if DEBUG :
                                print(f"> 4 > Token : {KEYCLOAK_TOKEN["return"]}")
                        print(f"> 4 >             > TOKEN KC renouvellé ")
                        
                    if ACTION == "CREATION" :
                        
                        if YOUDOC_CREATE_GROUPE :
                            print(f"> 4 > {GROUPE_NAME} > creation groupe Youdoc")
                            YOUDOC_ADMIN_GROUPE = YOUDOC_CONNEXION.admin_group(GROUPE_NAME,ACTION)
                            if YOUDOC_ADMIN_GROUPE["error"] :
                                print(f"> 4 > {GROUPE_NAME} > {YOUDOC_ADMIN_GROUPE["return"]}")
                                print("Appuyer sur entrer pour continuer")
                                input()
                            else :
                                print(f"> 4 > {GROUPE_NAME} > {YOUDOC_ADMIN_GROUPE["return"]}")
                        
                        if KEYCLOAK_CREATE_ROLE :                    
                            print(f"> 4 > {GROUPE_NAME} > creation role keycloak")
                            KEYCLOAK_ADMIN_ROLE = KEYCLOAK_CONNECT.admin_role(ROLE_NAME,ACTION)
                            if KEYCLOAK_ADMIN_ROLE["error"] :
                                print(f"> 4 > {GROUPE_NAME} > {KEYCLOAK_ADMIN_ROLE["return"]}")
                                print("Appuyer sur entrer pour continuer")
                                input()
                            else :
                                print(f"> 4 > {GROUPE_NAME} > {KEYCLOAK_ADMIN_ROLE["return"]}")
                            
                        
                        
                        
                        # MON_ID_ROLE = KEYCLOAK_CONNECT.id_role(ROLE_NAME)
                        # if MON_ID_ROLE["error"] :
                            # print(MON_ID_ROLE["return"])
                        # else :
                            # if DEBUG : print(f"{MON_ID_ROLE["return"]}")
                            # MON_ID_ROLE = MON_ID_ROLE["return"]
                        # print(MON_ID_ROLE)
                        
                        
                        if CREATE_MAPPER :
                            print(f"> 4 > {GROUPE_NAME} > creation mapper keycloak")
                            KEYCLOAK_ADMIN_MAPPER = KEYCLOAK_CONNECT.admin_mapper(IDP_ALIAS,CALCUL_CLIENT,GROUPE_NAME,ROLE_NAME,MAPPER_ID,"",ACTION)
                            if KEYCLOAK_ADMIN_MAPPER["error"] :
                                print(KEYCLOAK_ADMIN_MAPPER["return"])
                            else :
                                if DEBUG : print(f"{KEYCLOAK_ADMIN_MAPPER["return"]}")
                                print(f"> 4 > {GROUPE_NAME} > {KEYCLOAK_ADMIN_MAPPER["return"]}")
                            
                    if ACTION == "SUPPRESSION" :
                        if YOUDOC_CREATE_GROUPE :
                            print(f"> 4 > {GROUPE_NAME} > suppression groupe Youdoc")
                            YOUDOC_ADMIN_GROUPE = YOUDOC_CONNEXION.admin_group(GROUPE_NAME,ACTION)
                            if YOUDOC_ADMIN_GROUPE["error"] :
                                print(f"> 4 > {GROUPE_NAME} > {YOUDOC_ADMIN_GROUPE["return"]}")
                                print("Appuyer sur entrer pour continuer")
                                input()
                            else :
                                print(f"> 4 > {GROUPE_NAME} > {YOUDOC_ADMIN_GROUPE["return"]}")
                                
                            
                        
                        if CREATE_MAPPER :
                            print(f"> 4 > {GROUPE_NAME} > suppression mapper keycloak")
                            KEYCLOAK_ID_MAPPER = KEYCLOAK_CONNECT.id_mapper(IDP_ALIAS,GROUPE_NAME)
                            if KEYCLOAK_ID_MAPPER["error"] :
                                print(KEYCLOAK_ID_MAPPER["return"])
                            else :
                                if DEBUG : print(f"{KEYCLOAK_ID_MAPPER["return"]}")
                                KEYCLOAK_ID_MAPPER = KEYCLOAK_ID_MAPPER["return"]
                            
                                KEYCLOAK_ADMIN_MAPPER = KEYCLOAK_CONNECT.admin_mapper(IDP_ALIAS,CALCUL_CLIENT,GROUPE_NAME,ROLE_NAME,MAPPER_ID,KEYCLOAK_ID_MAPPER,ACTION)
                                if KEYCLOAK_ADMIN_MAPPER["error"] :
                                    print(KEYCLOAK_ADMIN_MAPPER["return"])
                                else :
                                    if DEBUG : print(f"{KEYCLOAK_ADMIN_MAPPER["return"]}")
                                    print(f"> 4 > {GROUPE_NAME} > {KEYCLOAK_ADMIN_MAPPER["return"]}")
                                
                        if KEYCLOAK_CREATE_ROLE :
                            print(f"> 4 > {GROUPE_NAME} > suppression role keycloak")
                            KEYCLOAK_ADMIN_ROLE = KEYCLOAK_CONNECT.admin_role(ROLE_NAME,ACTION)
                            if KEYCLOAK_ADMIN_ROLE["error"] :
                                print(f"> 4 > {GROUPE_NAME} > {KEYCLOAK_ADMIN_ROLE["return"]}")
                                print("Appuyer sur entrer pour continuer")
                                input()
                            else :
                                print(f"> 4 > {GROUPE_NAME} > {KEYCLOAK_ADMIN_ROLE["return"]}")
                                
                    print("")
                except Exception as e:
                    print("Appuyer sur entrer pour fermer")
                    print (f"> 4 > {GROUPE_NAME} > erreur {e}")
                    
    print(f"> 4 > Fin du traitement")                    

    if GENERATION_RAPPORT :
        print(f"> 5 > = Generation du rapport")  
        timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
        filename = Path.cwd() / f"Rapport_{timestamp}.csv"
        
        
        ligne_texte = "TYPE;LABEL;EXISTE;REALM;CLIENT;ROLE;GROUPE;INFO" + "\n"
        with filename.open(mode="a", encoding="utf-8") as f:
            f.write(ligne_texte)
            
        KEYCLOAK_TOKEN_EXPIRE = KEYCLOAK_CONNECT.token_valide()
        if KEYCLOAK_TOKEN_EXPIRE["error"] :
            print(KEYCLOAK_TOKEN_EXPIRE["return"])
            KEYCLOAK_TOKEN_EXPIRE_token = KEYCLOAK_TOKEN_EXPIRE["return"]
        else :
            KEYCLOAK_TOKEN_EXPIRE_token = KEYCLOAK_TOKEN_EXPIRE["return"]
        if DEBUG : print(f"> 5 > Token KC : {KEYCLOAK_TOKEN_EXPIRE}")
        if KEYCLOAK_TOKEN_EXPIRE_token :
            print(f"> 5 >             > TOKEN KC expiré ")
            KEYCLOAK_TOKEN = KEYCLOAK_CONNECT.recup_barear()
            if KEYCLOAK_TOKEN["error"] :
                print(KEYCLOAK_TOKEN["return"])
                print("Appuyer sur entrer pour fermer")
                input()
                sys.exit(1)
            else :
                if DEBUG :
                    print(f"> 5 > Token : {KEYCLOAK_TOKEN["return"]}")
            print(f"> 5 >             > TOKEN KC renouvellé ")
            
        # verification REALM
        print(f"> 5 > = Generation du rapport de REALM") 
        KEYCLOAK_VERIF_REALM = KEYCLOAK_CONNECT.verif_REALM()
        if KEYCLOAK_VERIF_REALM["error"] :
            ligne_texte = f"REALM;{CALCUL_CLIENT_REALM};NON;{CALCUL_CLIENT_REALM};;;;" + "\n"
        else :
            ligne_texte = f"REALM;{CALCUL_CLIENT_REALM};OUI;{CALCUL_CLIENT_REALM};;;;" + "\n"
        
        with filename.open(mode="a", encoding="utf-8") as f:
            f.write(ligne_texte)
        
        # verification CLIENT
        print(f"> 5 > = Generation du rapport de CLIENT") 
        KEYCLOAK_VERIF_CLIENT = KEYCLOAK_CONNECT.verif_CLIENT(CALCUL_CLIENT)
        if KEYCLOAK_VERIF_CLIENT["error"] :
            ligne_texte = f"CLIENT;{CALCUL_CLIENT};NON;{CALCUL_CLIENT_REALM};{CALCUL_CLIENT};;;" + "\n"
        else :
            ligne_texte = f"CLIENT;{CALCUL_CLIENT};OUI;{CALCUL_CLIENT_REALM};{CALCUL_CLIENT};;;" + "\n"
        
        with filename.open(mode="a", encoding="utf-8") as f:
            f.write(ligne_texte)
            
        # verification IDP
        if GENERATION_RAPPORT_MAPPER :

            print(f"> 5 > = Generation du rapport de IDP") 
            KEYCLOAK_VERIF_IDP = KEYCLOAK_CONNECT.verif_IDP(IDP_ALIAS)
            if KEYCLOAK_VERIF_IDP["error"] :
                ligne_texte = f"IDP;{IDP_ALIAS};NON;{CALCUL_CLIENT_REALM};;;;" + "\n"
            else :
                ligne_texte = f"IDP;{IDP_ALIAS};OUI;{CALCUL_CLIENT_REALM};;;;" + "\n"
            
            with filename.open(mode="a", encoding="utf-8") as f:
                f.write(ligne_texte)
                
        
        # verification GROUPES
        if GENERATION_RAPPORT_GROUPE :
            print(f"> 5 > = Generation du rapport des GROUPES") 
            
            YOUDOC_TOKEN_EXPIRE = YOUDOC_CONNEXION.token_valide()
            if YOUDOC_TOKEN_EXPIRE["error"] :
                print(YOUDOC_TOKEN_EXPIRE["return"])
                YOUDOC_TOKEN_EXPIRE_token = YOUDOC_TOKEN_EXPIRE["return"]
            else :
                YOUDOC_TOKEN_EXPIRE_token = YOUDOC_TOKEN_EXPIRE["return"]
            if DEBUG : print(f"> 5 > Token YD : {YOUDOC_TOKEN_EXPIRE_token}")
            if YOUDOC_TOKEN_EXPIRE_token :
                print(f"> 5 >             > TOKEN YD expiré ")
                YOUDOC_TOKEN = YOUDOC_CONNEXION.recup_barear()
                if YOUDOC_TOKEN["error"] :
                    print(YOUDOC_TOKEN["return"])
                    print("Appuyer sur entrer pour fermer")
                    input()
                    sys.exit(1)
                else :
                    if DEBUG :
                        print(f"> 5 > Token : {YOUDOC_TOKEN["return"]}")
                print(f"> 5 >             > TOKEN YD renouvellé ")
                
            YOUDOC_VERIF_group = YOUDOC_CONNEXION.list_group()
            if YOUDOC_VERIF_group["error"] :
                ligne_texte = f"GROUP;;ERROR;;;;;{YOUDOC_VERIF_group["return"]}" + "\n"
            else :
                YOUDOC_VERIF_group = YOUDOC_VERIF_group["return"]
                lst = ast.literal_eval(YOUDOC_VERIF_group)
                for item in lst:
                    
                    ligne_texte = f"GROUPE;{item.get("name")};OUI;;;;{item.get("name")};" + "\n"
                    with filename.open(mode="a", encoding="utf-8") as f:
                        f.write(ligne_texte)
            
        # verification ROLES
        if GENERATION_RAPPORT_ROLE :
            print(f"> 5 > = Generation du rapport des ROLES") 
            KEYCLOAK_TOKEN_EXPIRE = KEYCLOAK_CONNECT.token_valide()
            if KEYCLOAK_TOKEN_EXPIRE["error"] :
                print(KEYCLOAK_TOKEN_EXPIRE["return"])
                KEYCLOAK_TOKEN_EXPIRE_token = KEYCLOAK_TOKEN_EXPIRE["return"]
            else :
                KEYCLOAK_TOKEN_EXPIRE_token = KEYCLOAK_TOKEN_EXPIRE["return"]
            if DEBUG : print(f"> 4 > Token KC : {KEYCLOAK_TOKEN_EXPIRE}")
            if KEYCLOAK_TOKEN_EXPIRE_token :
                print(f"> 4 >             > TOKEN KC expiré ")
                KEYCLOAK_TOKEN = KEYCLOAK_CONNECT.recup_barear()
                if KEYCLOAK_TOKEN["error"] :
                    print(KEYCLOAK_TOKEN["return"])
                    print("Appuyer sur entrer pour fermer")
                    input()
                    sys.exit(1)
                else :
                    if DEBUG :
                        print(f"> 4 > Token : {KEYCLOAK_TOKEN["return"]}")
                print(f"> 4 >             > TOKEN KC renouvellé ")
            
            KEYCLOAK_VERIF_role = KEYCLOAK_CONNECT.liste_role()
            if KEYCLOAK_VERIF_role["error"] :
                ligne_texte = f"ROLE;;ERROR;{CALCUL_CLIENT_REALM};{CALCUL_CLIENT};;;{KEYCLOAK_VERIF_role["return"]}" + "\n"
            else :
                KEYCLOAK_VERIF_role = KEYCLOAK_VERIF_role["return"]
                lst = ast.literal_eval(KEYCLOAK_VERIF_role)
                for item in lst:
                    traitement = item.get("name").split('_', 2)[-1]
                    ligne_texte = f"ROLE;{item.get("name")};OUI;{CALCUL_CLIENT_REALM};{CALCUL_CLIENT};{CALCUL_CLIENT}.{item.get("name")};{traitement};" + "\n"
                    with filename.open(mode="a", encoding="utf-8") as f:
                        f.write(ligne_texte)
        
        # verification MAPPERS
        if GENERATION_RAPPORT_MAPPER :
            print(f"> 5 > = Generation du rapport des MAPPERS") 
            KEYCLOAK_TOKEN_EXPIRE = KEYCLOAK_CONNECT.token_valide()
            if KEYCLOAK_TOKEN_EXPIRE["error"] :
                print(KEYCLOAK_TOKEN_EXPIRE["return"])
                KEYCLOAK_TOKEN_EXPIRE_token = KEYCLOAK_TOKEN_EXPIRE["return"]
            else :
                KEYCLOAK_TOKEN_EXPIRE_token = KEYCLOAK_TOKEN_EXPIRE["return"]
            if DEBUG : print(f"> 4 > Token KC : {KEYCLOAK_TOKEN_EXPIRE}")
            if KEYCLOAK_TOKEN_EXPIRE_token :
                print(f"> 4 >             > TOKEN KC expiré ")
                KEYCLOAK_TOKEN = KEYCLOAK_CONNECT.recup_barear()
                if KEYCLOAK_TOKEN["error"] :
                    print(KEYCLOAK_TOKEN["return"])
                    print("Appuyer sur entrer pour fermer")
                    input()
                    sys.exit(1)
                else :
                    if DEBUG :
                        print(f"> 4 > Token : {KEYCLOAK_TOKEN["return"]}")
                print(f"> 4 >             > TOKEN KC renouvellé ")
            
            KEYCLOAK_VERIF_mapper = KEYCLOAK_CONNECT.liste_mapper(IDP_ALIAS)
            if KEYCLOAK_VERIF_mapper["error"] :
                ligne_texte = f"MAPPERS;;ERROR;{CALCUL_CLIENT_REALM};{CALCUL_CLIENT};;;{KEYCLOAK_VERIF_mapper["return"]}" + "\n"
            else :
                KEYCLOAK_VERIF_mapper = KEYCLOAK_VERIF_mapper["return"]
                lst = ast.literal_eval(KEYCLOAK_VERIF_mapper)
                for item in lst:
                    ligne_texte = f"MAPPERS;{item.get("name")};OUI;{CALCUL_CLIENT_REALM};{CALCUL_CLIENT};{item.get("config", {}).get("role")};{item.get("name")};{item.get("config", {}).get("claim.value")}" + "\n"
                    with filename.open(mode="a", encoding="utf-8") as f:
                        f.write(ligne_texte)
            
    print("")
    print("Bon courage pour la suite")
    print("")
    print(r"  (\_/)")
    print(r" =(^.^)=")
    print("  (LUVA)")
    print(" ( )_( )")
    print("")
    print("Appuyer sur entrer pour fermer")
    input()
    